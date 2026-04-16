import io
import asyncio
from datetime import date
from aiogram import Router, types, Bot
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from filters.admin_filter import IsAdminFilter
from states.states import AdminState
from keyboards.inline import get_admin_menu_keyboard, get_main_menu_keyboard
from utils.helpers import is_affirmative, get_user_id_or_username
from config import SUBSCRIPTION_PRICE, USER_DB_PATH

router = Router()
router.message.filter(IsAdminFilter())
router.callback_query.filter(IsAdminFilter())


# ============ АДМИН-ПАНЕЛЬ ============

@router.message(Command("admin"))
async def cmd_admin_panel(message: types.Message):
    await message.answer(
        "👑 Админ-панель\n\nВыберите действие:",
        reply_markup=get_admin_menu_keyboard()
    )


@router.callback_query(lambda c: c.data == "admin_users")
async def admin_users_callback(callback: types.CallbackQuery, user_db):
    users = user_db.get_all_users()
    if not users:
        await callback.message.edit_text("👥 Нет пользователей", reply_markup=get_admin_menu_keyboard())
        await callback.answer()
        return
    
    text = "👥 Список пользователей:\n\n"
    for u in users[:20]:  # Показываем первые 20
        text += f"🆔 ID: {u['user_id']}\n"
        text += f"👤 Имя: {u['first_name']}\n"
        if u['username']:
            text += f"📱 @{u['username']}\n"
        text += f"📅 Регистрация: {u['created_at'][:10]}\n"
        
        if u.get('is_forever'):
            text += "💳 Подписка: бессрочная\n"
        elif u.get('paid_until'):
            text += f"💳 Оплачено до: {u['paid_until']}\n"
        elif u.get('trial_end'):
            text += f"🎁 Триал до: {u['trial_end']}\n"
        text += "─" * 20 + "\n"
    
    if len(users) > 20:
        text += f"\n... и ещё {len(users) - 20} пользователей"
    
    await callback.message.edit_text(text, reply_markup=get_admin_menu_keyboard())
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_add_user")
async def admin_add_user_callback(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text(
        "➕ Введите ID пользователя или @username:"
    )
    await state.set_state(AdminState.waiting_for_user_id)
    await callback.answer()


@router.message(AdminState.waiting_for_user_id)
async def process_admin_user_id(message: types.Message, state: FSMContext, user_db):
    user_input = message.text.strip()
    user_id = await get_user_id_or_username(user_db, user_input)
    
    if not user_id:
        await message.answer(f"❌ Пользователь {user_input} не найден")
        await state.clear()
        return
    
    await state.update_data(user_id=user_id)
    await message.answer(
        f"✅ Найден пользователь ID: {user_id}\n\n"
        "Выберите тип подписки:\n\n"
        "1️⃣ - Навсегда (бессрочно)\n"
        "2️⃣ - На количество дней\n\n"
        "Отправьте 1 или 2"
    )
    await state.set_state(AdminState.waiting_for_days)


@router.message(AdminState.waiting_for_days)
async def process_admin_days(message: types.Message, state: FSMContext, user_db, bot: Bot):
    data = await state.get_data()
    user_id = data.get("user_id")
    choice = message.text.strip()
    
    user_info = user_db.get_user_info(user_id)
    user_name = user_info.get('first_name', f"ID {user_id}") if user_info else f"ID {user_id}"
    
    if choice == "1":
        user_db.activate_forever_subscription(user_id)
        await message.answer(
            f"✅ Пользователю {user_name} выдана бессрочная подписка!",
            reply_markup=get_admin_menu_keyboard()
        )
        await state.clear()
        
        try:
            await bot.send_message(
                user_id,
                "🎉 Вам выдана бессрочная подписка! Теперь вы можете пользоваться ботом без ограничений."
            )
        except:
            await message.answer("⚠️ Не удалось отправить уведомление пользователю")
            
    elif choice == "2":
        await message.answer("📅 Введите количество дней (например: 30)")
        await state.set_state(AdminState.waiting_for_days_value)
    else:
        await message.answer("❌ Пожалуйста, введите 1 или 2")


@router.message(AdminState.waiting_for_days_value)
async def process_admin_days_value(message: types.Message, state: FSMContext, user_db, bot: Bot):
    data = await state.get_data()
    user_id = data.get("user_id")
    
    try:
        days = int(message.text.strip())
        user_db.activate_subscription(user_id, days)
        
        user_info = user_db.get_user_info(user_id)
        user_name = user_info.get('first_name', f"ID {user_id}") if user_info else f"ID {user_id}"
        
        await message.answer(
            f"✅ Пользователю {user_name} выдана подписка на {days} дней!",
            reply_markup=get_admin_menu_keyboard()
        )
        await state.clear()
        
        try:
            await bot.send_message(
                user_id,
                f"🎉 Вам выдана подписка на {days} дней!\n\nОсталось дней: {days}"
            )
        except:
            await message.answer("⚠️ Не удалось отправить уведомление пользователю")
            
    except ValueError:
        await message.answer("❌ Неверное количество дней. Введите число.")
        await state.clear()


@router.callback_query(lambda c: c.data == "admin_extend")
async def admin_extend_callback(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text(
        "⏱ Введите ID или @username пользователя и количество дней через пробел.\n"
        "Пример: @username 30"
    )
    await state.set_state(AdminState.waiting_for_days_value)
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_broadcast")
async def admin_broadcast_callback(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text(
        "📨 Введите текст рассылки.\n\n"
        "Отправьте сообщение, которое нужно разослать ВСЕМ пользователям."
    )
    await state.set_state(AdminState.waiting_for_broadcast_text)
    await callback.answer()


@router.message(AdminState.waiting_for_broadcast_text)
async def process_broadcast_text(message: types.Message, state: FSMContext, user_db):
    text = message.text
    users = user_db.get_all_user_ids()
    
    await state.update_data(broadcast_text=text, broadcast_users=users)
    
    await message.answer(
        f"📨 Подтвердите рассылку.\n\n"
        f"👥 Получателей: {len(users)}\n"
        f"📝 Текст: {text[:200]}{'...' if len(text) > 200 else ''}\n\n"
        f"Отправьте ДА для подтверждения или НЕТ для отмены"
    )
    await state.set_state(AdminState.waiting_for_broadcast)


@router.message(AdminState.waiting_for_broadcast)
async def process_broadcast_confirm(message: types.Message, state: FSMContext, bot: Bot):
    if is_affirmative(message.text.lower()):
        data = await state.get_data()
        text = data.get("broadcast_text")
        users = data.get("broadcast_users")
        
        status_msg = await message.answer(f"📤 Начинаю рассылку для {len(users)} пользователей...")
        
        success = 0
        failed = 0
        
        for user_id in users:
            try:
                await bot.send_message(user_id, text)
                success += 1
            except Exception as e:
                failed += 1
            
            await asyncio.sleep(0.05)
        
        await status_msg.edit_text(
            f"✅ Рассылка завершена.\n"
            f"✅ Успешно: {success}\n"
            f"❌ Ошибок: {failed}",
            reply_markup=get_admin_menu_keyboard()
        )
        await state.clear()
    else:
        await message.answer("❌ Рассылка отменена", reply_markup=get_admin_menu_keyboard())
        await state.clear()


@router.callback_query(lambda c: c.data == "admin_backup")
async def admin_backup_callback(callback: types.CallbackQuery, bot: Bot):
    try:
        with open(USER_DB_PATH, 'rb') as f:
            file_data = f.read()
        
        await callback.message.answer_document(
            document=BufferedInputFile(file_data, filename="users.db"),
            caption="💾 Бэкап базы данных users.db"
        )
        await callback.answer("✅ Бэкап отправлен")
    except Exception as e:
        await callback.message.answer(f"❌ Ошибка: {e}")
        await callback.answer("❌ Ошибка")


@router.callback_query(lambda c: c.data == "admin_export")
async def admin_export_callback(callback: types.CallbackQuery, user_db):
    users = user_db.get_all_users_detailed()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    # Заголовки
    headers = ["ID", "Username", "Имя", "Дата регистрации", "Подписка до", 
               "Дней осталось", "Бессрочная", "Привёл рефералов", "Комиссия"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Данные
    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=user["user_id"])
        ws.cell(row=row_idx, column=2, value=user.get("username", ""))
        ws.cell(row=row_idx, column=3, value=user.get("first_name", ""))
        ws.cell(row=row_idx, column=4, value=user.get("created_at", "")[:10])
        ws.cell(row=row_idx, column=5, value=user.get("paid_until", ""))
        ws.cell(row=row_idx, column=6, value=user.get("days_left", 0))
        ws.cell(row=row_idx, column=7, value="Да" if user.get("is_forever") else "Нет")
        ws.cell(row=row_idx, column=8, value=user.get("total_refs", 0))
        ws.cell(row=row_idx, column=9, value=user.get("total_commission", 0))
    
    # Автоширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    today = date.today().strftime("%Y-%m-%d")
    await callback.message.answer_document(
        document=BufferedInputFile(excel_file.read(), filename=f"users_{today}.xlsx"),
        caption=f"📊 Список пользователей ({len(users)} чел.)"
    )
    await callback.answer("✅ Экспорт выполнен")


@router.callback_query(lambda c: c.data == "admin_refs")
async def admin_refs_callback(callback: types.CallbackQuery):
    from handlers.referral import cmd_ref_stats
    await cmd_ref_stats(callback.message, callback.bot.user_db)
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_stats")
async def admin_stats_callback(callback: types.CallbackQuery, user_db):
    users = user_db.get_all_user_ids()
    active_users = user_db.get_active_user_ids()
    
    text = (
        f"📊 Статистика бота:\n\n"
        f"👥 Всего пользователей: {len(users)}\n"
        f"✅ Активных подписок: {len(active_users)}\n"
        f"📈 Конверсия: {len(active_users)/len(users)*100 if users else 0:.1f}%"
    )
    
    await callback.message.edit_text(text, reply_markup=get_admin_menu_keyboard())
    await callback.answer()